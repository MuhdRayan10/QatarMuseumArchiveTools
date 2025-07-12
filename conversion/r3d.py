import subprocess
import threading
import ffmpeg
import time
import os

from openpyxl import Workbook, load_workbook


excel_file = "./conversion_log.xlsx"
excel_lock = threading.Lock()

def add_log_entry(input_path, output_path):
    with excel_lock:
        if not os.path.exists(excel_file):
            wb = Workbook()
            ws = wb.active
            ws.append(["Input Path", "Output Path"])
            wb.save(excel_file)

        wb = load_workbook(excel_file)
        ws = wb.active
        ws.append([input_path, output_path])
        wb.save(excel_file)

def r3d_to_mov(input_path, output_dir, resolution):
    output_path = os.path.join(output_dir, os.path.splitext(os.path.basename(input_path))[0] + '.mov')
    command = ["redline", "--exportPreset", "A", "-i", input_path, "-w", "201", "-R", str(resolution), 
               "--useRMD", "1", "-c", "1", "-G", "32", "--o", os.path.splitext(output_path)[0]]

    print("Running R3D -> MOV conversion...")
    subprocess.run(command, check=True)
    print(f"Conversion complete: {output_path}")

    return output_path

def mov_to_mp4(input_path):
    output_path = os.path.splitext(input_path)[0] + '.mp4'

    print("Running MOV -> MP4 conversion...")
    ffmpeg.input(input_path).output(output_path, vcodec='libx264', acodec='aac', preset="fast").run(overwrite_output=True)
    print(f"Conversion complete: {output_path}")

    return output_path


def convert_file(input_path, output_dir, resolution="2", export_format="mp4"):
    os.makedirs(output_dir, exist_ok=True)
    mov_path = r3d_to_mov(input_path, output_dir, resolution)
    if export_format == "mp4":
        mp4_path = mov_to_mp4(mov_path)
        final_output = os.path.join(output_dir, os.path.basename(mp4_path))
        os.replace(mp4_path, final_output)
        os.remove(mov_path)
    else:  # export_format == "mov"
        final_output = os.path.join(output_dir, os.path.basename(mov_path))
        os.replace(mov_path, final_output)
    add_log_entry(input_path, final_output)
    return final_output

def convert_directory(input_dir):

    start_time = time.time()
    file_count, data_handled = 0, 0

    input_dir = input_dir.rstrip(os.sep)
    
    parent = os.path.dirname(input_dir)
    base = os.path.basename(input_dir)
    
    output_root = os.path.join(parent, base + '_converted')

    for root, _, files in os.walk(input_dir):
        rel = os.path.relpath(root, input_dir)
        target_folder = os.path.join(output_root, rel)
        for file in files:
            if not file.lower().endswith('.r3d'):
                continue
            src = os.path.join(root, file)
            print(f"Converting {src} to {target_folder}")
            convert_file(src, target_folder)

            file_count += 1
            data_handled += os.path.getsize(src)

    total_time = time.time() - start_time
    print(f"Total conversion time: {total_time} seconds")
    print(f"Total files converted: {file_count}")
    print(f"Total data handled: {data_handled / (1024 * 1024):.2f} MB")
    print(f"Rate of conversion: {data_handled / (total_time) / (1024 * 1024):.2f} MB/s")
